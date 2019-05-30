# Khanh Nghiem
# Jun 04, 2018
# A command line program that prompts the user to choose excel files
# Extract class names and student data
# Use the Welsh-Powell algorithm to find blocks of courses that should be
# scheduled together

# Graphics Modules
from tkinter import Tk
from tkinter import filedialog
# Modules to process Excel files
from openpyxl import *
from openpyxl.compat import *
from openpyxl.utils import *

# Open file dialog, return a list of file addresses
def openFiles():
    # Close the default Tk GUI window
    root = Tk()
    root.withdraw()

    # Allow user to choose multiple files
    files =  filedialog.askopenfilenames(initialdir = ".",title = "Select file",filetypes = (("excel files","*.xlsx"),("all files","*.*")))
    files = list(files)

    root.update()
    root.destroy()

    return files

# Reading one single Excel file
def readExcel(filename):
    wb = load_workbook(filename)
    ws = wb['Sheet1']
    courseName = ws['A8'].value+' '+ws['B8'].value

    studentList = []
    FIRSTNAMECOL, LASTNAMECOL = 6,7
    row = 8
    while (ws.cell(column=FIRSTNAMECOL, row=row).value):
        studentList.append(ws.cell(column=FIRSTNAMECOL, row=row).value+' '+ws.cell(column=LASTNAMECOL, row=row).value)
        row += 1
    return courseName, studentList

# Reading one single Text files
def readTxt(filename):
    f = open(filename, 'r')
    lines = f.readlines()
    labels = lines[0].strip().split('\t')

    for i in range(len(labels)):
        if('ID' in labels[i]):
            id = i
        elif('SUB' in labels[i]):
            sub = i
        elif ('COURSE' in labels[i]):
            num = i
    rec = lines[1].strip().split('\t')
    courseName = rec[sub]+' '+rec[num]
    studentList = []
    studentList.append(rec[id])

    if len(lines) < 3:
        return studentList

    for k in range(2, len(lines)):
        rec = lines[k].strip().split('\t')
        studentList.append(rec[id])

    return courseName, studentList


# Read the list of Excel files and return the enrollment data
def process(files):
    # a dictionary with courses as keys and students as values
    enrollment = {}

    for f in files:
        courseName, studentList = readExcel(f)
        enrollment[courseName] = studentList

    return enrollment

# Create a graph of enrollment data
def createGraph(enrollment):
    # get the list of courses
    courseList = list(enrollment.keys())

    # initiate a graph with each vertex is a course and
    # there is an edge between two courses if they have the same student
    graph = {}
    for course in courseList:
        graph[course]=[]

    # for each key(class) check every other key to find out if any students
    # belong to two or more keys
    for i in range(len(courseList)-1):
        for k in range(i+1, len(courseList)):
            c1 = courseList[i]
            c2 = courseList[k]
            for s1 in enrollment[c1]:
                for s2 in enrollment[c2]:
                    #check if the students being considered are the same student
                    #make sure that it isn't the exact same dictionary entry
                    #(same student and class)
                    if s1==s2 and c2 not in graph[c1]:
                        graph[c1].append(c2)
                        graph[c2].append(c1)
    return graph

# return the degree of each vertex
def getDegree(pair):
    return len(pair[1])

# Black-box Welsh-Powell solver
def Welsh_Powell(graph):
    # sorting vertices by descending degree
    graph_list = list(graph.items())
    graph_list.sort(key = getDegree, reverse = True)

    # get the set of vertices in order
    V = []
    for el in graph_list:
        V.append(el[0])

    # initiate result coloring dict, with vertices as keys and colors as values
    coloring = {}

    # initiate the number (as also the index) of colors
    count = 0

    # traverse the the list and color the graph
    for i in range(len(V)):
        v = V[i]
        #print(v)
        # if the vertex has not been colored
        if v not in coloring:
            # assign the first available color
            coloring[v] = count

            # increment the count of colors
            count += 1

            # traverse list V and assign the same color to non-neigbors of v
            for k in range(i+1, len(V)):
                v_prime = V[k]
                if (v_prime not in coloring) and (v_prime not in graph[v]):

                    # check if any neighbor of v_prime is of the same color as v
                    # assume there is none (which means we could still color v_prime the same as v)
                    flag = False

                    # get all neighbors of v_prime
                    neighbors = graph[v_prime]

                    for n in neighbors:
                        # if a neighbor of v_prime has the same color as v, flag = True
                        if (n in coloring) and (coloring[n] == coloring[v]):
                            flag = True
                    # if v_prime doesn't have any neighbor of the same color as v
                    # color v_prime as v
                    if flag == False:
                        coloring[v_prime] = coloring[v]
    return coloring

def schedule(graph):
    free = []
    V = list(graph.keys())
    for v in V:
        if len(graph[v])==0:
            free.append(v)
            graph.pop(v)
    coloring = Welsh_Powell(graph)
    return coloring, free

def getColor(pair):
    return pair[1]

def blocks(coloring):
    if coloring=={}:
        return {}
    coloring_list = list(coloring.items())
    coloring_list.sort(key = getColor)
    dictionary = {}
    c = coloring_list[0][1]
    for el in coloring_list:
        if el[1] != c:
            c = el[1]
        if c not in dictionary:
            dictionary[c] = []
        dictionary[c].append(el[0])
    return dictionary



def main():
    files = openFiles()
    enrollment = process(files)
    graph = createGraph(enrollment)
    coloring, free = schedule(graph)
    print('Coloring:\n',coloring)
    print('Free courses:\n',free)
    print(blocks(coloring))

#if __name__ == "__main__":
#    main()
