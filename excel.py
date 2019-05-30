from openpyxl import *
from openpyxl.compat import *
from openpyxl.utils import *

def readExcel(filename):
    wb = load_workbook(filename)
    ws = wb['Sheet1']
    courseName = ws['A8'].value+' '+ws['B8'].value

    studentList = []
    row = 8
    while (ws.cell(column=6, row=row).value):
        studentList.append(ws.cell(column=6, row=row).value+' '+ws.cell(column=7, row=row).value)
        row += 1
    return courseName, studentList

def default():
    wb = Workbook()
    dest_filename = 'empty_book.xlsx'

    ws1 = wb.active
    ws1.title = "range names"

    for row in range(1, 40):
        ws1.append(range(600))

    ws2 = wb.create_sheet(title="Pi")

    ws2['F5'] = 3.14

    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

    print(ws3['AA10'].value)
    wb.save(filename = dest_filename)

def main():
    filename = 'registrar-data/ECO 248- Fall 2017.xlsx'
    courseName, studentList = readExcel(filename)
    print(courseName, studentList)

if __name__ == "__main__":
    main()
